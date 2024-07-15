import numpy as np
from openpyxl import load_workbook

def calcular_impedancias(workbook,frecuencia):
    sheet = workbook['Z']
    warnings = []
    impedancia=[]
    # Iterar sobre todas las filas a partir de la segunda (asumiendo que la primera es la cabecera)
    for i in range(2, sheet.max_row + 1):
        R = sheet[f'D{i}'].value
        L = sheet[f'E{i}'].value
        C = sheet[f'F{i}'].value
        nodoi=sheet[f'A{i}'].value
        nodoj=sheet[f'B{i}'].value
        # Si todos los valores son nulos, se detiene el cálculo
        if R is None and L is None and C is None:
            warnings.append(f"Fila {i}: Todos los valores son nulos. Se detiene el cálculo.")
            break

        R = R if R is not None else 0
        L = L if L is not None else 0
        C = C if C is not None else 0

        try:
            if C == 0:
                Z = complex(R, 2 * np.pi*frecuencia * L*10**-3)  # Evitar la división por cero
            else:
                Z = complex(R,((2 * np.pi * L*frecuencia *10**-3) - (1 / (2 * np.pi *frecuencia* C*10**-6))))
        except ZeroDivisionError:
            warnings.append(f"Fila {i}: División por cero detectada en el cálculo de Z.")
            break
        impedancia.append([nodoi,nodoj,Z])
    # Guardar advertencias en una hoja separada
    if warnings:
        if 'Warnings' not in workbook.sheetnames:
            workbook.create_sheet('Warnings')
        warning_sheet = workbook['Warnings']
        for j, warning in enumerate(warnings, start=1):
            warning_sheet.cell(row=j, column=1, value=warning)
        workbook.save('data_io.xlsx')
        #raise ValueError("Se encontraron problemas en los datos de entrada. Verifique la hoja 'Warnings' para más detalles.")
    return impedancia
# Cargar el archivo Excel y llamar a la función
