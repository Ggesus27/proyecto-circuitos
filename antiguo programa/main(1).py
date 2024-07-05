from z_equivalente import Ze
from impedancia import Zr, Zl, Zc
from warningV import checkV, crashV
from warningI import checkI, crashI
from warningZ import checkZ, crashZ
from crash import crashProgram

from store import storeNodesVI, storeNodesZ, storeVI, storeZ, storeF
from write import writeSheetVI, writeSheetZ
from openpyxl import Workbook
from potencias import crear_datos,potencia_fuentes_voltaje,potencia_fuentes_corriente,encontrar_nodo,potencia_z,delta_p
import openpyxl
import math
import cmath
book = openpyxl.load_workbook('data_io.xlsx')
write = Workbook()

sheet = book['f_and_ouput']
sheetV = book['V_fuente']
sheetI = book['I_fuente']
sheetZ = book['Z']

output = sheet['B2'].value
out1 = write.create_sheet('V_fuente')
out2 = write.create_sheet('I_fuente')
out3 = write.create_sheet('Z')

f = sheet['B1'].value
w = 2 * math.pi * f

nodesv = [0,0,0]
nodesi = [0,0,0]
nodesz = [[0,0,0], [0,0,0]]

zv = [0,0,0]
zi = [0,0,0]
z = [0,0,0]

fv = [0,0,0]
fi = [0,0,0]

checkV(book, sheetV, sheetV.max_row)
crashV(book, sheetV, sheetV.max_row)

checkI(book, sheetI, sheetI.max_row)
crashI(book, sheetI, sheetI.max_row)

checkZ(book, sheetZ, sheetZ.max_row)
crashZ(book, sheetZ, sheetZ.max_row)

cv = crashV(book, sheetV, sheetV.max_row)
ci = crashI(book, sheetI, sheetI.max_row)
cz = crashZ(book, sheetZ, sheetZ.max_row)

if crashProgram(cv, ci, cz) == 1:
    print('ERROR(!): Existen argumentos invalidos en la hoja de datos')
    exit(1)



storeNodesVI(nodesv, sheetV, sheetV.max_row)
storeVI(zv, Ze, Zr, Zl, Zc, w, sheetV, sheetV.max_row)
storeF(fv, sheetV, w, sheetV.max_row)

storeNodesVI(nodesi, sheetI, sheetI.max_row)
storeVI(zi, Ze, Zr, Zl, Zc, w, sheetI, sheetI.max_row)
storeF(fi, sheetI, w, sheetI.max_row)

storeNodesZ(nodesz, sheetZ, sheetZ.max_row)
storeZ(z, Ze, Zr, Zl, Zc, w, sheetZ, sheetZ.max_row)

writeSheetVI(zv, nodesv, fv, output, write, out1)
writeSheetVI(zi, nodesi, fi, output, write, out2)
writeSheetZ(z, nodesz, output, write, out3)

book2=openpyxl.load_workbook[output]
v_fuente,v_nodos,i_fuente,z_nodos=crear_datos(book2,w)
potencia_fuentes_voltaje(book2,v_fuente,v_nodos)
potencia_fuentes_corriente(book2,i_fuente,v_nodos)
potencia_z(z_nodos,v_nodos,book2)
delta_p(book)