import openpyxl

def verificar_datos(workbook, data_sheet, warning_col):
    warnings = []
    for row in data_sheet.iter_rows(min_row = 2, values_only = True):
        if any(cell is None or cell == '' for cell in row):
            warnings.append("Hay celdas vacías en la fila {}".format(row[0]))

    if warnings:
        if 'Warnings' not in workbook.sheetnames:
            workbook.create_sheet('Warnings')
        sheet = workbook['Warnings']
        for i, warning in enumerate(warnings, start=1):
            sheet.cell(row=i, column=warning_col, value=warning)
        workbook.save('data_io_output.xlsx')
        raise ValueError("Se encontraron problemas en los datos de entrada. Verifique la hoja 'Warnings' para más detalles.")

def verificar_fuentes_y_impedancias(workbook):
    sheets_to_check = ['I_fuente', 'V_fuente', 'Z']  # Nombres de las hojas a verificar
    for sheet_name in sheets_to_check:
        data_sheet = workbook[sheet_name]
        if sheet_name == 'Z':
            warning_col = 3  # Columna C para Z
        else:
            warning_col = 2  # Columna B para I_fuente y V_fuente
        verificar_datos(workbook, data_sheet, warning_col)