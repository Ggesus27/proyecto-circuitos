import openpyxl
import impedancias
import vth_and_zth as th
from v_fuente import calcular_voltajes_fuente as cal_vol
from I_fuente import calcular_corrientes_fuente as cal_co
import alertas as al
import numpy as np
from f_and_output import leer_frecuencia_y_salida as lf
import potencias
def main():
    # Especificar la ruta relativa al archivo Excel
    workbook_path = 'data_io.xlsx'
    
    try:
        # Cargar archivo Excel
        workbook = openpyxl.load_workbook(workbook_path)
    except FileNotFoundError:
        print("Error: El archivo no se encuentra en la ruta especificada.")
        return
    except PermissionError:
        print("Error: No se tienen permisos para acceder al archivo.")
        return

    try:
        # Leer y verificar datos de entrada
        frecuencia, archivo_salida=lf(workbook)
        sheet=workbook['V_fuente']
        sheet2=workbook['I_fuente']
        sheet3=workbook['Z']
        if al.verificar_fuentes(sheet)==1 or al.verificar_fuentes(sheet2)==1 or al.verificar_impedancias(sheet3)==1:
           workbook.save(archivo_salida)
           print("Hay errores en los datos colocados, verificar el archivo caso 1 para mas detalles")
           exit(1)
        workbook.save(archivo_salida)
        workbook=openpyxl.load_workbook(archivo_salida)
        # Calcular impedancias y convertir a fasores
        Z = impedancias.calcular_impedancias(workbook,frecuencia)
        v_fuente = cal_vol(workbook, frecuencia)
        i_fuente = cal_co(workbook, frecuencia)
        # Calcular admitancia
        Y = []
        for x in Z:
            Y.append([x[0], x[1], 1 / x[2]])
        for x in v_fuente:
            Y.append([0,x[0],1/x[2]])
        for x in i_fuente:
            Y.append([0,x[0],1/x[2]])
        #calculando thevenin
        cant_nodos=th.Cantidad_nodos(Y)
        Matriz_admitancias=th.matriz_a(cant_nodos,Y)
        matriz_resultados=th.matriz_b(i_fuente,cant_nodos,v_fuente)
        Zth,Vth=th.thevenin(Matriz_admitancias,matriz_resultados,cant_nodos)
        th.guardar_thevenin(Zth,Vth,workbook,archivo_salida)
        
        book=openpyxl.load_workbook(archivo_salida)
        #calculando potencias
        p_t_z=potencias.Calcular_potencias_impedancias(Vth,Z,book,archivo_salida)
        book=openpyxl.load_workbook(archivo_salida)
        p_t_f=potencias.guardar_fuentes(potencias.Calcular_potencia_voltaje(Vth,v_fuente),potencias.calcular_potencia_corriente(Vth,i_fuente),book,archivo_salida)
        potencias.balance(p_t_f,p_t_z,book,archivo_salida)
        
    except ValueError as e:
        print(e)

if __name__ == '__main__':
    main()