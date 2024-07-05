from z_equivalente import Ze
from impedancia import Zr, Zl, Zc
from warningV import checkV, crashV
from warningI import checkI, crashI
from warningZ import checkZ, crashZ
from crash import crashProgram
from out import generateOutputVI, generateOutputZ, writeSheetVI, writeSheetZ, storeNodesVI, storeNodesZ, storeVI, storeZ
from openpyxl import Workbook
import openpyxl
import math
#esto es un comentario
book = openpyxl.load_workbook(r'C:\Users\pc\OneDrive\Escritorio\data_io.xlsx')
write = Workbook()
excel_input = openpyxl.load_workbook(r'C:\Users\pc\OneDrive\Escritorio\data_io.xlsx')

sheet = book['f_and_ouput']
sheetV = book['V_fuente']
sheetI = book['I_fuente']
sheetZ = book['Z']
sheetVth = book['VTH_AND_ZTH']
output_filename = str(sheet["B2"].value)

output = sheet['B2'].value
out1 = write.create_sheet('V_fuente')
out2 = write.create_sheet('I_fuente')
out3 = write.create_sheet('Z')

f = sheet['B1'].value
w = 2 * math.pi * f

nodesv = [0,0,0]
nodesi = [0,0,0]
nodesz = [[0,0,0], [0,0,0]]
zv = [0,0,0]
zi = [0,0,0]
z = [0,0,0]

checkV(book, sheetV, sheetV.max_row)
crashV(book, sheetV, sheetV.max_row)

checkI(book, sheetI, sheetI.max_row)
crashI(book, sheetI, sheetI.max_row)

checkZ(book, sheetZ, sheetZ.max_row)
crashZ(book, sheetZ, sheetZ.max_row)

cv = crashV(book, sheetV, sheetV.max_row)
ci = crashI(book, sheetI, sheetI.max_row)
cz = crashZ(book, sheetZ, sheetZ.max_row)

if crashProgram(cv, ci, cz) == 1:
    print('ERROR(!): Existen argumentos invalidos en la hoja de datos')
    exit(1)

generateOutputVI(output, write, out1, out2)
generateOutputZ(output, write, out3)

storeNodesVI(nodesv, sheetV, sheetV.max_row)
storeVI(zv, Ze, Zr, Zl, Zc, w, sheetV, sheetV.max_row)

storeNodesVI(nodesi, sheetI, sheetI.max_row)
storeVI(zi, Ze, Zr, Zl, Zc, w, sheetI, sheetI.max_row)

storeNodesZ(nodesz, sheetZ, sheetZ.max_row)
storeZ(z, Ze, Zr, Zl, Zc, w, sheetZ, sheetZ.max_row)

writeSheetVI(zv, nodesv, output, write, out1)
writeSheetVI(zi, nodesi, output, write, out2)
writeSheetZ(z, nodesz, output, write, out3)

# calculate current/voltaje source's RMS value with phase current value
def calc_angle_img(number):  
  return math.degrees(math.atan2(number.imag, number.real))

def fasor(max_value, time_phase, frequency):
  
  # DC
  if (frequency == 0):    
    angle_phase = 0
    rms_value = max_value
  # AC  
  else:
    angle_phase = time_phase * ((2 * math.pi) / (1 / frequency))
    rms_value = max_value / math.sqrt(2)
  
  return (rms_value * math.cos(angle_phase)) + (rms_value * math.sin(angle_phase)) * (1j)



def z_equivalent(r, l, c, f):
  
  compl = 1j
  compc = -1j
  zl = 0
  zc = 0
  
  if (f == 0):
    
    # inductor
    
    # capacitor
    if (c == 0):
      zc = 0
    else:  
      zc = math.pow(10, 10)
    
  else:
    
    # inductor
    if (l == 0):
    #if l == type(float): 
      zl = 0
    elif l != None:  
      zl = (w * l * (10 ** -3))
    
    #capacitor
    if (zc == 0):
      zc = 0
    else:  

      zc = z = (1 / (w * c * (10 ** -6)))
  
  print(zl - zc)
  if zl != None and zc != None and r != None:
     return r + ((zl - zc)*(1j))
  # return math.sqrt(math.pow(resistance, 2) + math.pow(inductor_impedance - capacitor_impedance, 2))

def calc_impedance(resistance, inductance, capacitance, frequency):
  
  inductor_impedance = 0
  capacitor_impedance = 0
  
  if (frequency == 0):
    
    # inductor
    inductor_impedance = 0
    
    # capacitor
    if (capacitance == 0):
      capacitor_impedance = 0
    else:  
      capacitor_impedance = math.pow(10, 10)
    
  else:
    
    # inductor
    if (inductance == 0):
      inductor_impedance = 0
    else:  
      inductor_impedance = (1) * (2 * math.pi * frequency) * (inductance * math.pow(10, -3))
    
    #capacitor
    if (capacitance == 0):
      capacitor_impedance = 0
    else:  
      capacitor_impedance = (1) / ((2 * math.pi * frequency) * (capacitance * math.pow(10, -6)))
  
  print(inductor_impedance - capacitor_impedance)
  
  return resistance + ((inductor_impedance - capacitor_impedance) * (1j))
# contador de filas
row = 0 


for i in range(1, len(sheetV["A"])):
  if ((sheetV["E"][i].value == 0 and sheetV["F"][i].value == 0 and sheetV["Z"][i].value == 0)
      or
      (sheetV["E"][i].value == 0 and f == 0 and sheetV["Z"][i].value == 0)):
    sheetV["B"][i].value = "La impedancia total es 0 para la fuente de voltaje."
    print("La Impedancia en serie es 0 para la fuente de voltaje en la fila", i + 1)
    sheetV["E"][i].value = math.pow(10, -6)
    
  if (type(sheetV["A"][i].value) != int or sheetV["A"][i].value == 0):
    sheetV["B"][i].value = "Número de nodo de conexión inválido."
    print("Nodo de conexión no admitible.")
    excel_input.save(output_filename)
    exit()


# esta lista contiene todos los valores del circuito y los nodos a los que se conectan los elementos
circuit_data = []

# obtener voltajes
for i in range(len(sheetV["A"]) - 1):
  
  circuit_data.append([])
  circuit_data[row].append(0)
  circuit_data[row].append(int(sheetV["A"][i + 1].value)) 
  circuit_data[row].append(fasor(sheetV["C"][i + 1].value, sheetV["D"][i + 1].value, f))  # voltaje complejo
  circuit_data[row].append(None)
  circuit_data[row].append(z_equivalent( sheetV["E"][i + 1].value,sheetV["F"][i + 1].value,sheetV["Z"][i + 1].value,f)) # impedancia                        
  row = row + 1
  


# obtener corrientes
for i in range(len(I_fuente["A"]) - 1):
  
  circuit_data.append([])
  circuit_data[row].append(0) # siempre con respeto al nodo 0
  circuit_data[row].append(int(sheetI["A"][i + 1].value)) 
  circuit_data[row].append(None) 
  circuit_data[row].append(fasor( sheetI["C"][i + 1].value,sheetI["D"][i + 1].value, f))  # corriente compleja
  circuit_data[row].append(None)  
  # anexar la resistencia a otra columna  
  row = row + 1
  
  circuit_data.append([])
  circuit_data[row].append(0)
  circuit_data[row].append(int(sheetI["A"][i + 1].value)) 
  circuit_data[row].append(None) 
  circuit_data[row].append(None) 
  circuit_data[row].append(z_equivalent( sheetI["E"][i + 1].value,sheetI["F"][i + 1].value,sheetI["Z"][i + 1].value,f))
  row = row + 1


# impedancias sin fuentes en serie
for i in range(len(Z["A"]) - 1):
  
  circuit_data.append([])
   # nodo i
  circuit_data.append(int(Z["A"][i + 1].value)) 
  # nodo j
  circuit_data[row].append(int(Z["B"][i + 1].value))
  # no hay fuente de voltaje
  circuit_data[row].append(None) # no hay fuente de voltaje
  # no hay fuente de corriente
  circuit_data[row].append(None) 
  circuit_data[row].append(z_equivalent( Z["D"][i + 1].value, Z["E"][i + 1].value, Z["Z"][i + 1].value, f)) # impedancia
  
  row = row + 1


print()
for list in circuit_data:
  print(list)
print()  


# transformar fuentes de voltaje a corriente
# y añadirlas a circuit_data_list

# guardar copia
#aqui se guardan las fuentes de voltaje y corriente transformadas
circuit_data.copy = circuit_data.copy()

for i in range(len(circuit_data)):
  if (circuit_data[i][2] != None):
    
    # nueva fuente de corriente
    circuit_data.append([])
    circuit_data[len(circuit_data) - 1].append(circuit_data[i][0])
    circuit_data[len(circuit_data) - 1].append(circuit_data[i][1])
    circuit_data[len(circuit_data) - 1].append(None)
    circuit_data[len(circuit_data) - 1].append(circuit_data[i][2] / circuit_data[i][4])
    circuit_data[len(circuit_data) - 1].append(None)
    
    # impedancia nueva
    circuit_data.append([])
    circuit_data[len(circuit_data) - 1].append(circuit_data[i][0])
    circuit_data[len(circuit_data) - 1].append(circuit_data[i][1])
    circuit_data[len(circuit_data) - 1].append(None)
    circuit_data[len(circuit_data) - 1].append(None)
    circuit_data[len(circuit_data) - 1].append(circuit_data[i][4])    
    
    circuit_data[i] = []


print()
for list in circuit_data:
  print(list)
print()



max_value = -float('inf')


# Loop a traves de las columnas Bj y Bi para encontrar el mayor valor
for row in range(2, sheet3.max_row + 1):
    value = sheet3.cell(row=row, column=1).value
    if value is not None and value > max_value:
        max_value = value

    value = sheet3.cell(row=row, column=2).value
    if value is not None and value > max_value:
        max_value = value

# al final del loop, max value tiene el valor del nodo con mayor valor
last_node = max_value



# matriz de admitancia y el vector columna

ybus = numpy.zeros((last_node,last_node), dtype=complex)
column_vector = numpy.zeros((last_node,1), dtype=complex)

for i in range(1, last_node + 1):
  
  # lista que contendrá los elementos que se conectan a un nodo
  # el número de nodo analizandose lo da "i"
  node_elements = []
  
  for j in range(len(circuit_data)):
    
    if (circuit_data[j] == []):
      continue
    
    if (circuit_data[j][0] == i or circuit_data[j][1] == i):
      node_elements.append(circuit_data[j])
  
  print()
  for list in node_elements:
    print(list)
  print()
  
  # construir una fila del sistema de ecuaciones
  for j in range(len(node_elements)): 
    
    # si el elemento esta conctaado a i y 0
    if (node_elements[j][0] == 0 or node_elements[j][1] == 0):
      if (node_elements[j][3] != None): # fuente de corriente
        column_vector[i - 1][0] = column_vector[i - 1][0] + node_elements[j][3]
      else: # impedancia
        ybus[i - 1][i - 1] = ybus[i - 1][i - 1] + (1 / node_elements[j][4])
    
    # si el elemento esta conectado a i y j
    else:      
      temp_node = node_elements[j][0]
      if (i == temp_node): # the node number that is not i or 0
        temp_node = node_elements[j][1]
      
      ybus[i - 1][temp_node - 1] = ybus[i - 1][temp_node - 1] - (1 / node_elements[j][4])
      ybus[i - 1][i - 1] = ybus[i - 1][i - 1] + (1 / node_elements[j][4])


print(ybus)
print(column_vector)

inverse_ybus = numpy.linalg.inv(ybus)
node_voltages_matrix = inverse_ybus @ column_vector

print()
print(node_voltages_matrix)
print()


Sfuente = excel_input["Sfuente"]
S_Z = excel_input["S_Z"]

for i in range(1, last_row + 1):
  sheetVth["A"][i].value = i
  sheetVth["B"][i].value = abs(node_voltages_matrix[i - 1][0])
  sheetVth["C"][i].value = calc_angle_img(node_voltages_matrix[i - 1][0])
  sheetVth["D"][i].value = inverse_ybus[i - 1][i - 1].real
  sheetVth["E"][i].value = inverse_ybus[i - 1][i - 1].imag

  